"""缺勤统计 Excel 导出。"""

from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Dict, Iterable, Tuple


def build_absence_stats_filename(teaching_class_name: str = "") -> str:
    """构建缺勤统计导出文件名。"""
    safe_class_name = _sanitize_filename(teaching_class_name.strip())
    if safe_class_name:
        return f"{safe_class_name}-缺勤统计.xlsx"
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"缺勤统计-{timestamp}.xlsx"


def export_absence_stats_to_excel(
    absence_stats: Dict,
    total_activities: int,
    save_path: str,
    course_id: str = "",
    class_id: str = "",
    communication_status_getter=None,
) -> str:
    """导出缺勤统计到 Excel，返回最终保存路径。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "缺勤统计"

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A1:G1")
    worksheet["A1"] = f"缺勤统计导出（共 {total_activities} 次签到活动）"
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = center_alignment

    headers = ["姓名", "学号", "班级", "缺勤次数", "总签到次数", "缺勤率", "沟通情况"]
    worksheet.append(headers)

    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for row in _build_rows(
        absence_stats=absence_stats,
        total_activities=total_activities,
        course_id=course_id,
        class_id=class_id,
        communication_status_getter=communication_status_getter,
    ):
        worksheet.append(row)

    widths = {
        "A": 14,
        "B": 18,
        "C": 28,
        "D": 12,
        "E": 14,
        "F": 12,
        "G": 12,
    }
    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width

    for row in worksheet.iter_rows(min_row=3, min_col=4, max_col=7):
        for cell in row:
            cell.alignment = center_alignment

    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    workbook.save(save_path)
    return save_path


def _build_rows(
    absence_stats: Dict,
    total_activities: int,
    course_id: str,
    class_id: str,
    communication_status_getter,
) -> Iterable[Tuple[str, str, str, int, int, str, str]]:
    sorted_stats = sorted(
        absence_stats.items(),
        key=lambda item: item[1].get("absent_count", 0),
        reverse=True,
    )
    for uid, stats in sorted_stats:
        absent_count = int(stats.get("absent_count", 0))
        total_count = int(stats.get("total_count", total_activities))
        denominator = total_count or total_activities or 0
        absence_rate = f"{(absent_count / denominator * 100):.1f}%" if denominator else "0.0%"
        communicated = False
        if communication_status_getter:
            try:
                communicated = bool(communication_status_getter(course_id, class_id, int(uid)))
            except Exception:
                communicated = False

        yield (
            str(stats.get("name", "")),
            str(stats.get("username", "")),
            str(stats.get("class_name", "")),
            absent_count,
            total_count,
            absence_rate,
            "已沟通" if communicated else "未沟通",
        )


def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
