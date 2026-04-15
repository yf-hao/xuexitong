"""作业情况 Excel 导出。"""

from __future__ import annotations

import os
import re
from datetime import datetime
from typing import Iterable, List, Tuple


def build_homework_stats_filename(teaching_class_name: str = "") -> str:
    """构建作业情况导出文件名。"""
    safe_class_name = _sanitize_filename(teaching_class_name.strip())
    if safe_class_name:
        return f"{safe_class_name}-作业情况.xlsx"
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"作业情况-{timestamp}.xlsx"


def export_homework_stats_to_excel(
    stats_list: List,
    save_path: str,
    course_id: str = "",
    class_id: str = "",
    communication_status_getter=None,
) -> str:
    """导出作业情况到 Excel。"""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "作业情况"

    title_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    title_font = Font(color="FFFFFF", bold=True, size=12)
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")

    worksheet.merge_cells("A1:J1")
    worksheet["A1"] = f"作业情况导出（共 {len(stats_list)} 名学生）"
    worksheet["A1"].fill = title_fill
    worksheet["A1"].font = title_font
    worksheet["A1"].alignment = center_alignment

    headers = ["学号", "姓名", "作业数", "已提交", "待批", "未提交", "平均分", "最低分", "最高分", "沟通情况"]
    worksheet.append(headers)

    for cell in worksheet[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    for row in _build_rows(
        stats_list=stats_list,
        course_id=course_id,
        class_id=class_id,
        communication_status_getter=communication_status_getter,
    ):
        worksheet.append(row)

    widths = {
        "A": 18,
        "B": 14,
        "C": 10,
        "D": 10,
        "E": 10,
        "F": 10,
        "G": 10,
        "H": 10,
        "I": 10,
        "J": 12,
    }
    for col, width in widths.items():
        worksheet.column_dimensions[col].width = width

    for row in worksheet.iter_rows(min_row=3, min_col=3, max_col=10):
        for cell in row:
            cell.alignment = center_alignment

    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    workbook.save(save_path)
    return save_path


def _build_rows(
    stats_list: List,
    course_id: str,
    class_id: str,
    communication_status_getter,
) -> Iterable[Tuple[str, str, int, int, int, int, float, float, float, str]]:
    for stats in stats_list:
        communicated = False
        if communication_status_getter:
            try:
                communicated = bool(communication_status_getter(course_id, class_id, int(stats.person_id)))
            except Exception:
                communicated = False

        yield (
            str(stats.alias_name),
            str(stats.user_name),
            int(stats.complete_num),
            int(stats.work_submitted),
            int(stats.pending_count),
            int(stats.unsubmitted_count),
            round(float(stats.real_avg_score), 2),
            round(float(stats.min_score), 1),
            round(float(stats.max_score), 1),
            "已沟通" if communicated else "未沟通",
        )


def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]+', "_", name).strip()
